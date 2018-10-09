# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import os
import re
import sys
import time
import Tkinter as tk
from tkFileDialog import askopenfilename
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/util')
from utility import *

reload(sys)
sys.setdefaultencoding('utf8')

versao = "1.9.3.5"
if len(sys.argv) > 1:
	filename = sys.argv[1]
	if os.path.isfile(filename):
		path = os.path.dirname(filename)
	else:
		path = filename
		filename = os.path.join(path, 'Produtos.xlsx')
else:
	root = tk.Tk()
	root.withdraw()
	filename = askopenfilename()
	if not filename:
		print("Nenhum arquivo selecionado!")
		sys.exit(1)
	path = os.path.dirname(filename)

wb = load_workbook(filename, keep_vba=True)

# grab the active worksheet
ws = wb.active

# detect start of data
row = 1
col = 1
while True:
	value = ws.cell(column=col, row=row).value
	if value or col > 100:
		break;
	row += 1
	col += 1
col_first = col
col_row =row
rfirst = row + 1
# detect row count
col = rfirst - 1
row = rfirst
rlast = rfirst - 1
while True:
	value = ws.cell(column=col, row=row).value
	if value == None:
		break;
	row += 1
	rlast = row

# detect code columns
code_columns = find_columns(ws, r"^(?:codigo|cod|id|numero)$", col_first, col_row)
if not code_columns:
	print("Nenhuma coluna de codigo do produto encontrada!")
# detect barcode columns
barcode_columns = find_columns(ws, r"barra", col_first, col_row)
if not barcode_columns:
	print("Nenhuma coluna de codigo de barras do produto encontrada!")
# detect description columns
description_columns = find_columns(ws, r"descricao|nome|produto", col_first, col_row)
if not description_columns:
	print("Nenhuma coluna de descricao do produto encontrada!")
	sys.exit(1)
# detect price name columns
price_columns = find_columns(ws, r"valor|preco", col_first, col_row)
if not price_columns:
	print("Nenhuma coluna do preco do produto encontrada!")
	sys.exit(1)
# detect category columns
category_columns = find_columns(ws, r"categoria|grupo|classe", col_first, col_row)
if not category_columns:
	print("Nenhuma coluna de categoria encontrada!")
	sys.exit(1)
subcategory_columns = find_columns(ws, r"subcategoria|subgrupo|subclasse", col_first, col_row)
if not subcategory_columns:
	print("Nenhuma coluna de subcategoria encontrada!")
# detect details columns
details_columns = find_columns(ws, r"detalhe", col_first, col_row)
if not details_columns:
	print("Nenhuma coluna de detalhes encontrada!")
# detect abreviation columns
abreviation_columns = find_columns(ws, r"abreviacao|apelido", col_first, col_row)
if not abreviation_columns:
	print("Nenhuma coluna de abreviacao encontrada!")
# detect unity columns
unity_columns = find_columns(ws, r"unidade", col_first, col_row)
if not unity_columns:
	print("Nenhuma coluna de unidade encontrada!")
# detect sector columns
sector_columns = find_columns(ws, r"setor|impressao|impressora", col_first, col_row)
if not sector_columns:
	print("Nenhuma coluna de setor encontrada!")
# detect visibility columns
visibility_columns = find_columns(ws, r"visivel|ativo|mostrar", col_first, col_row)
if not visibility_columns:
	print("Nenhuma coluna de visibilidade encontrada!")
# detect divisibility columns
divisibility_columns = find_columns(ws, r"divisivel|fracionado", col_first, col_row)
if not divisibility_columns:
	print("Nenhuma coluna de divisibilidade encontrada!")
# detect type columns
type_columns = find_columns(ws, r"tipo", col_first, col_row)
if not type_columns:
	print("Nenhuma coluna de tipo encontrada!")
# detect quantity limit columns
limit_columns = find_columns(ws, r"limite", col_first, col_row)
if not limit_columns:
	print("Nenhuma coluna de quantidade limite encontrada!")
# detect stock control columns
stock_columns = find_columns(ws, r"control.*estoq", col_first, col_row)
if not stock_columns:
	print("Nenhuma coluna de controle de estoque encontrada!")
# detect stock control columns
service_columns = find_columns(ws, r"servico", col_first, col_row)
if not service_columns:
	print("Nenhuma coluna de cobrar servico encontrada!")
# detect ncm columns
ncm_columns = find_columns(ws, r"ncm", col_first, col_row)
if not ncm_columns:
	print("Nenhuma coluna de NCM encontrada!")
# detect imposto columns
imposto_columns = find_columns(ws, r"(?:imposto|ˆst$|ˆcst$)", col_first, col_row)
if not imposto_columns:
	print("Nenhuma coluna de imposto encontrada!")
# detect origem columns
origem_columns = find_columns(ws, r"origem", col_first, col_row)
if not origem_columns:
	print("Nenhuma coluna de origem encontrada!")
# detect cfop columns
cfop_columns = find_columns(ws, r"cfop", col_first, col_row)
if not cfop_columns:
	print("Nenhuma coluna de cfop encontrada!")
# detect cest columns
cest_columns = find_columns(ws, r"cest", col_first, col_row)
if not cest_columns:
	print("Nenhuma coluna de cest encontrada!")

# save data
filename = os.path.join(path, "MySQLBackup.utf8.sql")
with open(filename, "w") as fd:
	fd.write("/* Backup Version: GrandChef " + versao + " */\n")
	fd.write("/* Backup on " + time.strftime("%d-%m-%Y %H:%M:%S") + " */\n\n")
	fd.write("USE `GrandChef`;\n")

	fd.write("\n-- Categorias\n")
	categories = set()
	for row in xrange(rfirst, rlast):
		categoria = get_cell(ws, row, category_columns)
		if not categoria:
			continue
		categoria = filter_name(categoria)
		if categoria in categories:
			continue
		categories.add(categoria)
		servico = "Y"
		fd.write("INSERT INTO Categorias (CategoriaID, Descricao, Servico, DataAtualizacao) VALUES\n")
		fd.write("	(NULL, " + sql_field(categoria) + ", " + sql_field(servico, "'") + ", NOW()) ON DUPLICATE KEY UPDATE CategoriaID = VALUES(CategoriaID), DataAtualizacao = VALUES(DataAtualizacao);\n")


	fd.write("\n-- Subcategorias\n")
	subcategories = set()
	for row in xrange(rfirst, rlast):
		subcategoria = get_cell(ws, row, subcategory_columns)
		categoria = get_cell(ws, row, category_columns)
		if not categoria or not subcategoria or categoria == subcategoria:
			continue
		categoria = filter_name(categoria)
		subcategoria = filter_name(subcategoria)
		key = categoria + '.' + subcategoria
		if key in subcategories:
			continue
		subcategories.add(key)
		servico = "Y"
		fd.write("INSERT INTO Categorias (CategoriaID, Descricao, Servico, DataAtualizacao) VALUES\n")
		fd.write("	((SELECT c.ID FROM Categorias c WHERE c.Descricao = " + sql_field(categoria) + " AND ISNULL(c.CategoriaID)), " + sql_field(subcategoria) + ", " + sql_field(servico, "'") + ", NOW()) ON DUPLICATE KEY UPDATE CategoriaID = VALUES(CategoriaID), DataAtualizacao = VALUES(DataAtualizacao);\n")

	fd.write("\n-- Unidades\n")
	unitys = set()
	for row in xrange(rfirst, rlast):
		unidade = get_cell(ws, row, unity_columns)
		if not unidade:
			continue
		if unidade.lower() == "un":
			unidade = "UN"
		if unidade in unitys:
			continue
		unitys.add(unidade)
		nome = "Unidade " + unidade
		fd.write("INSERT INTO Unidades (Nome, Sigla) VALUES\n")
		fd.write("	(" + sql_field(nome) + ", " + sql_field(unidade, "'") + ") ON DUPLICATE KEY UPDATE Sigla = VALUES(Sigla);\n")

	fd.write("\n-- Setores\n")
	sectors = set()
	for row in xrange(rfirst, rlast):
		setor = get_cell(ws, row, sector_columns)
		if not setor:
			continue
		if setor in sectors:
			continue
		sectors.add(setor)
		descricao = "Setor " + setor
		fd.write("INSERT INTO Setores (Nome, Descricao) VALUES\n")
		fd.write("	(" + sql_field(setor) + ", " + sql_field(descricao, "'") + ") ON DUPLICATE KEY UPDATE Nome = VALUES(Nome);\n")

	fd.write("\n-- Tributacoes\n")
	taxation = set()
	for row in xrange(rfirst, rlast):
		codigo_produto = get_cell(ws, row, code_columns)
		if not codigo_produto:
			continue
		if codigo_produto in taxation:
			continue
		ncm = get_cell(ws, row, ncm_columns)
		if len(str(ncm)) != 8:
			ncm = None
		if not ncm: 
			continue
		cest = get_cell(ws, row, cest_columns)
		origem = get_cell(ws, row, origem_columns)
		if origem == None or len(str(origem)) == 0:
			continue
		operacao = get_cell(ws, row, cfop_columns)
		if not operacao:
			continue
		imposto = get_cell(ws, row, imposto_columns)
		if not imposto:
			continue
		taxation.add(codigo_produto)
		fd.write("INSERT INTO Tributacoes (ID, NCM, CEST, OrigemID, OperacaoID, ImpostoID) VALUES\n")
		fd.write("	(" + sql_int(codigo_produto) + ", " + sql_field(ncm) + ", " + sql_field(cest) + ", (SELECT ID FROM Origens WHERE Codigo = " + sql_int(origem) + "), (SELECT ID FROM Operacoes WHERE Codigo = " + sql_int(operacao) + "), (SELECT ID FROM Impostos WHERE Codigo = " + sql_int(imposto) + ")) ON DUPLICATE KEY UPDATE ID = VALUES(ID);\n")

	fd.write("\n-- Produtos\n")
	for row in xrange(rfirst, rlast):
		codigo = get_cell(ws, row, code_columns)
		if not codigo:
			codigo = None
		codigo_barras = get_cell(ws, row, barcode_columns)
		if not codigo_barras:
			codigo_barras = None
		descricao = get_cell(ws, row, description_columns)
		if not descricao:
			continue
		descricao = filter_name(descricao)
		categoria = get_cell(ws, row, category_columns)
		if not categoria:
			continue
		categoria = filter_name(categoria)
		subcategoria = get_cell(ws, row, subcategory_columns)
		subcategoria = filter_name(subcategoria)
		if not subcategoria or categoria == subcategoria:
			categoria_sql = "(SELECT ID FROM Categorias WHERE Descricao = " + sql_field(categoria) + ")"
		else:
			categoria_sql = "(SELECT s.ID FROM Categorias s LEFT JOIN Categorias c ON c.ID = s.CategoriaID WHERE s.Descricao = " + sql_field(subcategoria) + " AND c.Descricao = " + sql_field(categoria) + " AND NOT ISNULL(c.ID))"
		preco = get_cell(ws, row, price_columns)
		if not preco:
			preco = 0.00
		abreviacao = get_cell(ws, row, abreviation_columns)
		detalhes = get_cell(ws, row, details_columns)
		unidade = get_cell(ws, row, unity_columns)
		if not unidade:
			unidade = "UN"
		setor = get_cell(ws, row, sector_columns)
		if not setor and not sector_columns:
			setor = "Cozinha"
		controlar = get_cell(ws, row, stock_columns)
		tipo = get_cell(ws, row, type_columns)
		if not tipo and sql_bool(controlar, 'N') == "'Y'":
			tipo = "Produto"
		elif not tipo:
			tipo = "Composicao"
		limite = get_cell(ws, row, limit_columns)
		if not limite:
			limite = 0.00
		visivel = get_cell(ws, row, visibility_columns)
		divisivel = get_cell(ws, row, divisibility_columns)
		cobrar_servico = get_cell(ws, row, service_columns)
		tributacao = codigo if codigo in taxation else None
		fd.write("INSERT INTO Produtos (ID, CodigoBarras, Descricao, PrecoVenda, Abreviacao, Detalhes, Tipo, QuantidadeLimite, Visivel, Divisivel, CobrarServico, DataAtualizacao, CategoriaID, UnidadeID, SetorPreparoID, TributacaoID) VALUES\n")
		fd.write("	(" + sql_int(codigo) + ", " + sql_field(codigo_barras) + ", " + sql_field(descricao) + ", " + sql_float(preco) + ", " + sql_field(abreviacao) + ", " + sql_field(detalhes) + ", " + sql_field(tipo, "'") + ", " + sql_float(limite) + ", " + sql_bool(visivel, 'Y') + ", " + sql_bool(divisivel, 'N') + ", " + sql_bool(cobrar_servico, 'Y') + ", NOW(), " +  categoria_sql + ", (SELECT ID FROM Unidades WHERE Sigla = " + sql_field(unidade) + "), (SELECT ID FROM Setores WHERE Nome = " + sql_field(setor) + "), (" + sql_int(tributacao) + ")) ON DUPLICATE KEY UPDATE DataAtualizacao = VALUES(DataAtualizacao);\n")

	fd.close()
	convert_encoding(filename, os.path.join(path, "MySQLBackup.sql"))
	os.remove(filename)