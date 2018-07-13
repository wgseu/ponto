# -*- coding: utf-8 -*-
import os
import re
import sys
import time
import datetime
import json
import codecs
import roman
import MySQLdb
import Tkinter as tk
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from tkFileDialog import askopenfilename
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/util')
from utility import *


reload(sys)
sys.setdefaultencoding('utf-8')

versao = "1.9.3.0"

if len(sys.argv) > 1:
	filename = sys.argv[1]
	if os.path.isfile(filename):
		path = os.path.dirname(filename)
	else:
		path = filename
		filename = os.path.join(path, 'Clientes.xlsx')
	data = json.load(open(os.path.join(path, 'default.json')))
	if data['filename']:
		filename = os.path.join(path, data['filename'])
	ddd = str(data['ddd'])
	estado_padrao = data['estado']
	cidade_padrao = data['cidade']
	if len(sys.argv) > 2:
		ddd = sys.argv[2]
	if len(sys.argv) > 3:
		estado_padrao = sys.argv[3]
	if len(sys.argv) > 4:
		cidade_padrao = sys.argv[4]
else:
	root = tk.Tk()
	root.withdraw()
	filename = askopenfilename()
	if not filename:
		print("Nenhum arquivo selecionado!")
		sys.exit(1)
	path = os.path.dirname(filename)

wb = load_workbook(filename, keep_vba=True)

# grab the active worksheet
ws = wb.active

# detect start of data
row = 1
col = 1
while True:
	value = ws.cell(column=col, row=row).value
	if value or col > 100:
		break;
	row += 1
	col += 1
col_first = col
col_row =row
rfirst = row + 1
# detect row count
col = rfirst - 1
row = rfirst
rlast = rfirst - 1
while True:
	value = ws.cell(column=col, row=row).value
	if value == None:
		break;
	row += 1
	rlast = row

# detect phone columns
phone_columns = find_columns(ws, r"telefone|fone|celular", col_first, col_row)
if not phone_columns:
	print("Nenhuma coluna de telefone encontrada!")
	sys.exit(1)
# detect costumer name columns
nome_columns = find_columns(ws, r"nome|cliente", col_first, col_row)
if not nome_columns:
	print("Nenhuma coluna do nome do cliente encontrada!")
	sys.exit(1)
# detect E-mail columns
email_columns = find_columns(ws, r"email|e-mail", col_first, col_row)
if not email_columns:
	print("Nenhuma coluna de E-mail encontrada!")
# detect gender columns
gender_columns = find_columns(ws, r"sexo|genero|masculino|feminino|homem|mulher", col_first, col_row)
if not gender_columns:
	print("Nenhuma coluna de genero encontrada!")
# detect CPF columns
cpf_columns = find_columns(ws, r"CPF|documento", col_first, col_row)
if not cpf_columns:
	print("Nenhuma coluna de CPF encontrada!")
# detect RG/IE columns
rg_columns = find_columns(ws, r"(?:^rg|^ie|rg$|ie$|^registro.*geral$|identidade)", col_first, col_row)
if not rg_columns:
	print("Nenhuma coluna de RG/IE encontrada!")
# detect credit limit columns
credit_columns = find_columns(ws, r"limite|credito", col_first, col_row)
if not credit_columns:
	print("Nenhuma coluna de limite de credito encontrada!")
# detect register date columns
date_columns = find_columns(ws, r"(?:data.*)?(?:cadastro|registro)", col_first, col_row)
if not date_columns:
	print("Nenhuma coluna de data de cadastro encontrada!")
# detect register date columns
birthday_columns = find_columns(ws, r"(?:data.*)?aniversario|(?:data.*)?nascimento", col_first, col_row)
if not birthday_columns:
	print("Nenhuma coluna de data de aniversário encontrada!")
# detect address name columns
address_columns = find_columns(ws, r"endereco|rua|logradouro", col_first, col_row)
if not address_columns:
	print("Nenhuma coluna de logradouro encontrada!")
# detect neighborhood columns
neighborhood_columns = find_columns(ws, r"bairro", col_first, col_row)
if not neighborhood_columns:
	print("Nenhuma coluna de bairro encontrada!")
# detect tax columns
tax_columns = find_columns(ws, r"taxa|entrega", col_first, col_row)
if not tax_columns:
	print("Nenhuma coluna de valor de entrega encontrada!")
# detect city columns
city_columns = find_columns(ws, r"cidade|municipio", col_first, col_row)
if not city_columns:
	print("Nenhuma coluna de cidade encontrada!")
# detect state or state code columns
state_columns = find_columns(ws, r"estado", col_first, col_row)
state_code_columns = find_columns(ws, r"UF", col_first, col_row)
if not state_code_columns and not state_columns:
	print("Nenhuma coluna de estado encontrada!")
# detect postal code columns
zipcode_columns = find_columns(ws, r"CEP", col_first, col_row)
if not zipcode_columns:
	print("Nenhuma coluna de CEP encontrada!")
# detect number columns
number_columns = find_columns(ws, r"numero|casa", col_first, col_row)
if not number_columns:
	print("Nenhuma coluna de numero da casa encontrada!")
# detect condominio columns
condominio_columns = find_columns(ws, r"condominio", col_first, col_row)
if not condominio_columns:
	print("Nenhuma coluna de nome do condominio encontrada!")
# detect apartment columns
apartment_columns = find_columns(ws, r"apartamento|^ap$|^apt$|^apto$", col_first, col_row)
if not apartment_columns:
	print("Nenhuma coluna de numero do apartamento encontrada!")
# detect block columns
block_columns = find_columns(ws, r"bloco", col_first, col_row)
if not block_columns:
	print("Nenhuma coluna do bloco do apartamento encontrada!")
# detect reference columns
observation_columns = find_columns(ws, r"observacoes|observacao|referencia", col_first, col_row)
if not observation_columns:
	print("Nenhuma coluna de observacao ou referencia encontrada!")
# detect complement columns
complement_columns = find_columns(ws, r"complemento", col_first, col_row)
if not complement_columns:
	print("Nenhuma coluna de complemento encontrada!")

# save data
filename = os.path.join(path, "MySQLBackup.utf8.sql")
with open(filename, "w") as fd:
	fd.write("/* Backup Version: GrandChef " + versao + " */\n")
	fd.write("/* Backup on " + time.strftime("%d-%m-%Y %H:%M:%S") + " */\n\n")
	fd.write("USE `GrandChef`;\n")
	fd.write("\n-- Clientes\n")
	ids = {}
	for row in xrange(rfirst, rlast):
		phones = []
		for col in phone_columns:
			value = get_cell(ws, row, [col])
			phone = filter_phone(value, ddd)
			if phone and not phone in phones:
				phones.append(phone)
		if not phones:
			continue
		nome = get_cell(ws, row, nome_columns)
		if not nome:
			continue
		sobrenome = None
		cpf = get_cell(ws, row, cpf_columns)
		cpf = filter_cpf(cpf) or filter_cnpj(cpf)
		genero = get_cell(ws, row, gender_columns)
		tipo = 'Fisica'
		if genero and genero.lower() == 'empresa':
			tipo = 'Juridica'
		if filter_cnpj(cpf):
			tipo = 'Juridica'
		if tipo == 'Fisica':
			components = filter_name(nome).split(' ', 1)
			nome = components[0]
			if len(components) > 1:
				sobrenome = components[1]

		genero = filter_gender(genero)
		if not genero:
			genero = detect_gender(nome)
		fone1 = phones[0]
		fone2 = None
		if len(phones) > 1:
			fone2 = phones[1]
		data_cadastro = get_cell(ws, row, date_columns)
		data_aniversario = get_cell(ws, row, birthday_columns)
		rg = get_cell(ws, row, rg_columns)
		rg = filter_digits(rg)
		email = get_cell(ws, row, email_columns)
		email = filter_email(email)
		limite = get_cell(ws, row, credit_columns)
		ids[row] = fone1
		fd.write("INSERT INTO Clientes (Nome, Sobrenome, Tipo, Genero, CPF, RG, Email, Fone1, Fone2, LimiteCompra, DataAniversario, DataAtualizacao, DataCadastro) VALUES\n")
		fd.write("	(" + sql_field(nome) + ", " + sql_field(sobrenome) + ", " + sql_field(tipo, "'") + ", " + sql_field(genero, "'") + ", " + sql_field(cpf) + ", " + sql_field(rg) + ", " + sql_field(email) + ", " + sql_field(fone1) + ", " + sql_field(fone2) + ", " +
		         sql_float(limite) + ", " + sql_datetime(data_aniversario, 'NULL') + ", NOW(), " + sql_datetime(data_cadastro) + ") ON DUPLICATE KEY UPDATE Nome = VALUES(Nome), Sobrenome = VALUES(Sobrenome), Fone2 = VALUES(Fone2);\n")

	fd.write("\n-- Cidades\n")
	cidades = set()
	for row, fone1 in ids.iteritems():
		nome = get_cell(ws, row, city_columns)
		nome = filter_name(nome)
		if not nome and cidade_padrao in cidades:
			continue
		if not nome:
			nome = cidade_padrao
		uf = get_cell(ws, row, state_code_columns)
		estado = get_cell(ws, row, state_columns)
		if estado and len(estado) == 2:
			uf = estado
		if not estado and not uf:
			estado = estado_padrao
		if uf and len(uf) == 2:
			cond = "UF = " + sql_field(uf)
			cidade_key = nome + " - " + uf
		else:
			cond = "Nome = " + sql_field(estado)
			cidade_key = nome + " - " + estado
		if cidade_key in cidades:
			continue
		cidades.add(cidade_key)
		fd.write("INSERT INTO Cidades (Nome, EstadoID) VALUES\n")
		fd.write("	(" + sql_field(nome) + ", (SELECT ID FROM Estados WHERE " + cond + ")) ON DUPLICATE KEY UPDATE Nome = VALUES(Nome);\n")

	fd.write("\n-- Bairros\n")
	bairros = set()
	for row, fone1 in ids.iteritems():
		nome = get_cell(ws, row, neighborhood_columns)
		nome = filter_name(nome)
		if not nome:
			continue
		cidade = get_cell(ws, row, city_columns)
		cidade = filter_name(cidade)
		if not cidade:
			cidade = cidade_padrao
		bairro_key = nome + " - " + cidade
		if bairro_key in bairros:
			continue
		bairros.add(bairro_key)
		taxa = get_cell(ws, row, tax_columns)
		if not taxa:
			taxa = 0.00
		fd.write("INSERT INTO Bairros (Nome, ValorEntrega, CidadeID) VALUES\n")
		fd.write("	(" + sql_field(nome) + ", " + sql_float(taxa) + ", (SELECT ID FROM Cidades WHERE Nome = " + sql_field(cidade) + ")) ON DUPLICATE KEY UPDATE Nome = VALUES(Nome);\n")

	fd.write("\n-- Localizacoes\n")
	r_ini = r'^[\-, ]*'
	r_end = r'[\-, ]*$'
	r_sep = r'[\-, ]+'
	r_logra = r_ini + r'(.*?)'
	r_numer = r'(?:N)?[\.\- º°,]*([0-9]+[^\- ,]*?)'
	r_apart = r'(?:APARTAMENTO|APART|APRT|APTO|APT0|APT|AP)[\.\- ,]*([0-9]+[^\- ,]*?)'
	r_bloco = r'(?:BLOCO|BLOC|BLO|BLC|BL|B)[\.\- ,]*([^\- ,]+?)'
	r_edifi = r'(?:EDIFICIO|EDF|ED)[\.\- ,]+(.*?)'
	r_ed_op = r'(?:EDIFICIO|EDF|ED)?[\.\- ,]?(.*?)'
	r_casa = r'[\-, ]*(?:CASA)?'

	# 'R Marechal Manoel Luiz Ozorio, N 375, Ap 302, BL A, Edf Ravena'
	pnabe = re.compile(r_logra + r_sep + r_numer + r_sep + r_apart + r_sep + r_bloco + r_sep + r_edifi + r_end, re.I)
	# 'R Marechal Manoel Luiz Ozorio, N 375, BL A, Ap 302, Edf Ravena'
	pnbae = re.compile(r_logra + r_sep + r_numer + r_sep + r_bloco + r_sep + r_apart + r_sep + r_ed_op + r_end, re.I)
	# 'R Marechal Manoel Luiz Ozorio, N 375, BL A, Edf Ravena, Ap 302'
	pnbea = re.compile(r_logra + r_sep + r_numer + r_sep + r_bloco + r_sep + r_edifi + r_sep + r_apart + r_end, re.I)
	# 'R Marechal Manoel Luiz Ozorio, N 375, Ap 302, Edf Ravena, BL A'
	pnaeb = re.compile(r_logra + r_sep + r_numer + r_sep + r_apart + r_sep + r_edifi + r_sep + r_bloco + r_end, re.I)
	# 'R Marechal Manoel Luiz Ozorio, Edf Ravena, N 375, BL A, Ap 302'
	pneba = re.compile(r_logra + r_sep + r_numer + r_sep + r_edifi + r_sep + r_bloco + r_sep + r_apart + r_end, re.I)
	# 'R Marechal Manoel Luiz Ozorio, Edf Ravena, N 375, Ap 302, BL A'
	pneab = re.compile(r_logra + r_sep + r_numer + r_sep + r_edifi + r_sep + r_apart + r_sep + r_bloco + r_end, re.I)
	# 'R Marechal Manoel Luiz Ozorio, N 375, Ap 302 Edf Ravena'
	pnae = re.compile(r_logra + r_sep + r_numer + r_sep + r_apart + r_sep + r_ed_op + r_end, re.I)
	# 'R Marechal Manoel Luiz Ozorio, N 375, BL A, Edf Ravena'
	pnbe = re.compile(r_logra + r_sep + r_numer + r_sep + r_bloco + r_sep + r_edifi + r_end, re.I)
	# 'R Marechal Manoel Luiz Ozorio, N 375 Edf Ravena, Ap 302'
	pnea = re.compile(r_logra + r_sep + r_numer + r_sep + r_edifi + r_sep + r_apart + r_end, re.I)
	# 'R Marechal Manoel Luiz Ozorio, N 375, Edf Ravena, BL A'
	pneb = re.compile(r_logra + r_sep + r_numer + r_sep + r_edifi + r_sep + r_bloco + r_end, re.I)
	# 'AV MARIO ALVARES N 905 BLOC 8 AP 201'
	pnba = re.compile(r_logra + r_sep + r_numer + r_sep + r_bloco + r_sep + r_apart + r_end, re.I)
	# 'AV PROF MORAES N788 APT 302 BLOCO B'
	pnab = re.compile(r_logra + r_sep + r_numer + r_sep + r_apart + r_sep + r_bloco + r_end, re.I)
	# 'AV PROF MORAES N788 APT 302'
	pna = re.compile(r_logra + r_sep + r_numer + r_sep + r_apart + r_end, re.I)
	# 'AV PROF MORAES N788 BLOCO B'
	pnb = re.compile(r_logra + r_sep + r_numer + r_bloco + r_end, re.I)


	# 'RUA 6 BLOCO 53 AP 304'
	pba = re.compile(r_logra + r_sep + r_bloco + r_sep + r_apart + r_end, re.I)
	# 'RUA 6 AP 304 BLOCO 53'
	pab = re.compile(r_logra + r_sep + r_apart + r_sep + r_bloco + r_end, re.I)
	# 'RUA PIAUI, N 364'
	pn = re.compile(r_logra + r_sep + r_numer + r_casa + r_end, re.I)
	# 'RUA PIAUI, 364'
	pn2 = re.compile(r_logra + r_sep + r_numer + r_casa + r_end, re.I)


	# 'BLOC 8 AP 201'
	poba = re.compile(r_ini + r_bloco + r_sep + r_apart + r_end, re.I)
	# 'APT 302 BLOCO B'
	poab = re.compile(r_ini + r_apart + r_sep + r_bloco + r_end, re.I)

	# 'AP 201'
	poa = re.compile(r_ini + r_apart + r_end, re.I)

	for row, fone1 in ids.iteritems():
		bairro = get_cell(ws, row, neighborhood_columns)
		bairro = filter_name(bairro)
		if not bairro:
			continue
		cidade = get_cell(ws, row, city_columns)
		cidade = filter_name(cidade)
		if not cidade:
			cidade = cidade_padrao
		cep = get_cell(ws, row, zipcode_columns)
		cep = filter_zipcode(cep)
		logradouro = get_cell(ws, row, address_columns)
		logradouro = filter_name(logradouro)
		if not logradouro:
			continue
		numero = get_cell(ws, row, number_columns)
		apartamento = get_cell(ws, row, apartment_columns)
		bloco = get_cell(ws, row, block_columns)
		condominio = get_cell(ws, row, condominio_columns)
		referencia = get_cell(ws, row, observation_columns)
		referencia = filter_name(referencia)
		complemento = get_cell(ws, row, complement_columns)
		complemento = filter_name(complemento)
		if apartamento:
			tipo = 'Apartamento'
		else:
			tipo = 'Casa'
		apelido = 'Minha Casa'

		if not numero:
			m = pnabe.search(logradouro)
			if m:
				# print sql_field(logradouro) + "\t\t\t = LOG: " + sql_field(m.group(1)) + " + N: " + sql_field(m.group(2)) + " + BL: " + sql_field(m.group(4)) + " + APT: " + sql_field(m.group(3)) + " + COND: " + sql_field(m.group(5))
				logradouro = m.group(1)
				numero = m.group(2)
				bloco = filter_name(m.group(4))
				apartamento = m.group(3)
				condominio = filter_name(m.group(5))
				tipo = 'Apartamento'
		if not numero:
			m = pnbea.search(logradouro)
			if m:
				# print sql_field(logradouro) + "\t\t\t = LOG: " + sql_field(m.group(1)) + " + N: " + sql_field(m.group(2)) + " + BL: " + sql_field(m.group(4)) + " + APT: " + sql_field(m.group(3)) + " + COND: " + sql_field(m.group(5))
				logradouro = m.group(1)
				numero = m.group(2)
				bloco = filter_name(m.group(3))
				apartamento = m.group(5)
				condominio = filter_name(m.group(4))
				tipo = 'Apartamento'
		if not numero:
			m = pnaeb.search(logradouro)
			if m:
				# print sql_field(logradouro) + "\t\t\t = LOG: " + sql_field(m.group(1)) + " + N: " + sql_field(m.group(2)) + " + BL: " + sql_field(m.group(3)) + " + APT: " + sql_field(m.group(4)) + " + COND: " + sql_field(m.group(5))
				logradouro = m.group(1)
				numero = m.group(2)
				bloco = filter_name(m.group(5))
				apartamento = m.group(3)
				condominio = condominio or filter_name(m.group(4))
				tipo = 'Apartamento'
		if not numero:
			m = pneba.search(logradouro)
			if m:
				# print sql_field(logradouro) + "\t\t\t = LOG: " + sql_field(m.group(1)) + " + N: " + sql_field(m.group(2)) + " + BL: " + sql_field(m.group(4)) + " + APT: " + sql_field(m.group(5)) + " + COND: " + sql_field(m.group(3))
				logradouro = m.group(1)
				numero = m.group(2)
				bloco = filter_name(m.group(4))
				apartamento = m.group(5)
				condominio = condominio or filter_name(m.group(3))
				tipo = 'Apartamento'
		if not numero:
			m = pneab.search(logradouro)
			if m:
				# print sql_field(logradouro) + "\t\t\t = LOG: " + sql_field(m.group(1)) + " + N: " + sql_field(m.group(2)) + " + BL: " + sql_field(m.group(5)) + " + APT: " + sql_field(m.group(4)) + " + COND: " + sql_field(m.group(3))
				logradouro = m.group(1)
				numero = m.group(2)
				bloco = filter_name(m.group(5))
				apartamento = m.group(4)
				condominio = condominio or filter_name(m.group(3))
				tipo = 'Apartamento'
		if not numero:
			m = pnbae.search(logradouro)
			if m:
				# print sql_field(logradouro) + "\t\t\t = LOG: " + sql_field(m.group(1)) + " + N: " + sql_field(m.group(2)) + " + BL: " + sql_field(m.group(3)) + " + APT: " + sql_field(m.group(4)) + " + COND: " + sql_field(m.group(5))
				logradouro = m.group(1)
				numero = m.group(2)
				bloco = filter_name(m.group(3))
				apartamento = m.group(4)
				condominio = condominio or filter_name(m.group(5))
				tipo = 'Apartamento'
		if not numero:
			m = pnbe.search(logradouro)
			if m:
				# print sql_field(logradouro) + "\t\t\t = LOG: " + sql_field(m.group(1)) + " + N: " + sql_field(m.group(2)) + " + BL: " + sql_field(m.group(3)) + " + COND: " + sql_field(m.group(4))
				logradouro = m.group(1)
				numero = m.group(2)
				bloco = filter_name(m.group(3))
				condominio = condominio or filter_name(m.group(4))
				tipo = 'Apartamento'
		if not numero:
			m = pnea.search(logradouro)
			if m:
				# print sql_field(logradouro) + "\t\t\t = LOG: " + sql_field(m.group(1)) + " + N: " + sql_field(m.group(2)) + " + APT: " + sql_field(m.group(4)) + " + COND: " + sql_field(m.group(3))
				logradouro = m.group(1)
				numero = m.group(2)
				apartamento = m.group(4)
				condominio = condominio or filter_name(m.group(3))
				tipo = 'Apartamento'
		if not numero:
			m = pneb.search(logradouro)
			if m:
				# print sql_field(logradouro) + "\t\t\t = LOG: " + sql_field(m.group(1)) + " + N: " + sql_field(m.group(2)) + " + BL: " + sql_field(m.group(4)) + " + COND: " + sql_field(m.group(3))
				logradouro = m.group(1)
				numero = m.group(2)
				bloco = filter_name(m.group(4))
				condominio = condominio or filter_name(m.group(3))
				tipo = 'Apartamento'
		if not numero:
			m = pnba.search(logradouro)
			if m:
				# print sql_field(logradouro) + "\t\t\t = LOG: " + sql_field(m.group(1)) + " + N: " + sql_field(m.group(2)) + " + BL: " + sql_field(m.group(3)) + " + APT: " + sql_field(m.group(4)) 
				logradouro = m.group(1)
				numero = m.group(2)
				bloco = filter_name(m.group(3))
				apartamento = m.group(4)
				tipo = 'Apartamento'
		if not numero:
			m = pnab.search(logradouro)
			if m:
				# print sql_field(logradouro) + "\t\t\t = LOG: " + sql_field(m.group(1)) + " + N: " + sql_field(m.group(2)) + " + BL: " + sql_field(m.group(4)) + " + APT: " + sql_field(m.group(3)) 
				logradouro = m.group(1)
				numero = m.group(2)
				bloco = filter_name(m.group(4))
				apartamento = m.group(3)
				tipo = 'Apartamento'
		if not numero:
			m = pnae.search(logradouro)
			if m:
				# print sql_field(logradouro) + "\t\t\t = LOG: " + sql_field(m.group(1)) + " + N: " + sql_field(m.group(2)) + " + APT: " + sql_field(m.group(3)) + " + COND: " + sql_field(m.group(4))
				logradouro = m.group(1)
				numero = m.group(2)
				apartamento = m.group(3)
				condominio = condominio or filter_name(m.group(4))
				tipo = 'Apartamento'
		if not numero:
			m = pna.search(logradouro)
			if m:
				# print sql_field(logradouro) + "\t\t\t = LOG: " + sql_field(m.group(1)) + " + N: " + sql_field(m.group(2)) + " + APT: " + sql_field(m.group(3)) 
				logradouro = m.group(1)
				numero = m.group(2)
				apartamento = m.group(3)
				tipo = 'Apartamento'
		if not numero:
			m = pnb.search(logradouro)
			if m:
				# print sql_field(logradouro) + "\t\t\t = LOG: " + sql_field(m.group(1)) + " + N: " + sql_field(m.group(2)) + " + BL: " + sql_field(m.group(3))
				logradouro = m.group(1)
				numero = m.group(2)
				bloco = filter_name(m.group(3))
				tipo = 'Apartamento'
		if not numero:
			m = pn.search(logradouro)
			if m:
				# print sql_field(logradouro) + "\t\t\t = LOG: " + sql_field(m.group(1)) + " + N: " + sql_field(m.group(2)) 
				logradouro = m.group(1)
				numero = m.group(2)
		if not numero:
			m = pn2.search(logradouro)
			if m:
				# print sql_field(logradouro) + "\t\t\t = LOG: " + sql_field(m.group(1)) + " + N: " + sql_field(m.group(2)) 
				logradouro = m.group(1)
				numero = m.group(2)
		if not numero:
			numero = ''

		if complemento and not apartamento and not bloco:
			m = poab.search(complemento)
			if m:
				# print sql_field(logradouro) + "\t\t\t = BL: " + sql_field(m.group(2)) + " + APT: " + sql_field(m.group(1))
				apartamento = m.group(1)
				bloco = m.group(2)
				tipo = 'Apartamento'
				complemento = None
		if complemento and not apartamento and not bloco:
			m = poba.search(complemento)
			if m:
				# print sql_field(logradouro) + "\t\t\t = BL: " + sql_field(m.group(1)) + " + APT: " + sql_field(m.group(2))
				apartamento = m.group(2)
				bloco = m.group(1)
				tipo = 'Apartamento'
				complemento = None
		if complemento and not apartamento and not bloco:
			m = poa.search(complemento)
			if m:
				# print sql_field(logradouro) + "\t\t\t = APT: " + sql_field(m.group(1))
				apartamento = m.group(1)
				tipo = 'Apartamento'
				complemento = None
		if logradouro and len(logradouro) > 100:
			print('Logradouro "' + logradouro + '" truncado para o cliente: ' + fone1)
			logradouro = logradouro[:97] + '...'
		if complemento and len(complemento) > 100:
			print('Complemento "' + complemento + '" truncado para o cliente: ' + fone1)
			complemento = complemento[:97] + '...'
		if referencia and len(referencia) > 200:
			print('Referencia "' + referencia + '" truncada para o cliente: ' + fone1)
			referencia = referencia[:197] + '...'
		if numero and len(numero) > 20:
			print('Numero "' + numero + '" truncado para o cliente: ' + fone1)
			numero = numero[:20]

		fd.write("INSERT INTO Localizacoes (Apelido, CEP, Logradouro, Numero, Tipo, Condominio, Bloco, Apartamento, Complemento, Referencia, ClienteID, BairroID) VALUES\n")
		fd.write("	(" + sql_field(apelido) + ", " + sql_field(cep) + ", " + sql_field(logradouro) + ", " + sql_field(numero) + ", " + sql_field(tipo) + ", " + sql_field(condominio) + ", " + sql_field(bloco) + ", " + sql_field(apartamento) + ", " + sql_field(complemento) + ", " + sql_field(referencia) + ", (SELECT ID FROM Clientes WHERE Fone1 = " + sql_field(fone1) + "), (SELECT ID FROM Bairros WHERE Nome = " + sql_field(bairro) + " AND CidadeID = (SELECT ID FROM Cidades WHERE Nome = " + sql_field(cidade) + "))) ON DUPLICATE KEY UPDATE Tipo = VALUES(Tipo), Logradouro = VALUES(Logradouro), Numero = VALUES(Numero), Condominio = VALUES(Condominio), Bloco = VALUES(Bloco), Apartamento = VALUES(Apartamento), Complemento = VALUES(Complemento), Referencia = VALUES(Referencia);\n")

	fd.close()
	convert_encoding(filename, os.path.join(path, "MySQLBackup.sql"))
	os.remove(filename)
